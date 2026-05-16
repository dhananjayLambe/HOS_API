from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from diagnostics_engine.models.catalog import DiagnosticServiceMaster
from diagnostics_engine.models.choices import CommissionType
from diagnostics_engine.services.pricing_templates import constants as C
from diagnostics_engine.services.pricing_templates.excel_utils import (
    find_pricing_header_row,
    iter_table_rows,
    parse_bool,
    parse_decimal,
    parse_int,
    read_metadata_sheet,
    validate_import_metadata,
)
from diagnostics_engine.services.pricing_templates.validators import (
    validate_import_headers,
    validate_supported_pricing_row,
)
from labs.models import BranchServicePricing, LabBranch


class LabPricingImportStrictAbort(Exception):
    """First row failure in --strict mode; rolls back transaction."""

    def __init__(self, message: str, stats: ImportStats | None = None) -> None:
        super().__init__(message)
        self.message = message
        self.stats = stats


@dataclass
class ImportStats:
    """Import counters: skipped = not offered (FALSE/blank gate); unchanged = TRUE row, DB already matched."""

    created: int = 0
    updated: int = 0
    skipped: int = 0
    unchanged: int = 0
    failed: int = 0
    errors: List[str] = field(default_factory=list)


def _load_branch(meta: Dict[str, str], explicit_branch_code: str | None) -> LabBranch:
    code = (explicit_branch_code or meta.get("branch_code") or "").strip()
    if not code:
        raise ValueError("Metadata missing branch_code.")
    meta_code = (meta.get("branch_code") or "").strip()
    if explicit_branch_code and meta_code and explicit_branch_code != meta_code:
        raise ValueError(
            f"branch_code mismatch: metadata={meta_code!r}, argument={explicit_branch_code!r}"
        )
    try:
        return LabBranch.objects.select_related("organization").prefetch_related("address").get(
            branch_code=code,
            is_deleted=False,
        )
    except LabBranch.DoesNotExist as exc:
        raise ValueError(f"LabBranch not found for branch_code={code!r}") from exc


def _resolve_service(code: str) -> DiagnosticServiceMaster:
    return DiagnosticServiceMaster.objects.get(
        code=code,
        is_active=True,
        deleted_at__isnull=True,
    )


def _phase1_snapshots(selling_price: Decimal, cost_price: Decimal) -> tuple[Decimal, Decimal, Decimal]:
    """lab_payout, doctor_margin, platform_margin (Phase 1: no doctor slice)."""
    return cost_price, Decimal("0"), selling_price - cost_price


def _upsert_row(
    branch: LabBranch,
    svc: DiagnosticServiceMaster,
    *,
    selling_price: Decimal,
    cost_price: Decimal,
    report_hours: int,
    home_collection: bool,
    remarks: str,
    dry_run: bool,
) -> str:
    """Upsert active BranchServicePricing. Imported rows are always is_available=True."""
    existing = (
        BranchServicePricing.objects.filter(
            branch=branch,
            service=svc,
            is_deleted=False,
            is_active=True,
        )
        .order_by("-updated_at")
        .first()
    )

    lab_payout, doctor_snap, platform_snap = _phase1_snapshots(selling_price, cost_price)

    metadata: Dict[str, Any] = dict(existing.metadata) if existing and existing.metadata else {}
    if remarks:
        metadata["remarks"] = remarks
    elif "remarks" in metadata:
        metadata.pop("remarks")

    if dry_run:
        return "created" if existing is None else "updated"

    if existing is None:
        BranchServicePricing.objects.create(
            branch=branch,
            service=svc,
            selling_price=selling_price,
            cost_price=cost_price,
            platform_margin_type=CommissionType.FLAT,
            platform_margin_value=Decimal("0"),
            doctor_commission_type=CommissionType.FLAT,
            doctor_commission_value=Decimal("0"),
            valid_from=timezone.now().date(),
            is_active=True,
            is_available=True,
            home_collection_supported=home_collection,
            report_delivery_hours=report_hours,
            metadata=metadata,
            lab_payout_snapshot=lab_payout,
            doctor_margin_snapshot=doctor_snap,
            platform_margin_snapshot=platform_snap,
            currency="INR",
        )
        return "created"

    changed = False
    if existing.selling_price != selling_price:
        existing.selling_price = selling_price
        changed = True
    if existing.cost_price != cost_price:
        existing.cost_price = cost_price
        changed = True
    if existing.report_delivery_hours != report_hours:
        existing.report_delivery_hours = report_hours
        changed = True
    if existing.home_collection_supported != home_collection:
        existing.home_collection_supported = home_collection
        changed = True
    if existing.is_available is not True:
        existing.is_available = True
        changed = True
    if existing.metadata != metadata:
        existing.metadata = metadata
        changed = True
    if existing.lab_payout_snapshot != lab_payout:
        existing.lab_payout_snapshot = lab_payout
        changed = True
    if existing.doctor_margin_snapshot != doctor_snap:
        existing.doctor_margin_snapshot = doctor_snap
        changed = True
    if existing.platform_margin_snapshot != platform_snap:
        existing.platform_margin_snapshot = platform_snap
        changed = True

    if changed:
        existing.save()
        return "updated"
    return "skipped"


def import_lab_pricing(
    path: Path,
    *,
    branch_code: str | None = None,
    dry_run: bool = False,
    strict: bool = False,
) -> ImportStats:
    stats = ImportStats()
    wb = load_workbook(filename=str(path), data_only=True)
    try:
        if C.METADATA_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {C.METADATA_SHEET_NAME!r}")
        if C.PRICING_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {C.PRICING_SHEET_NAME!r}")

        meta = read_metadata_sheet(wb[C.METADATA_SHEET_NAME])
        validate_import_metadata(meta)
        branch = _load_branch(meta, branch_code)

        ws_price = wb[C.PRICING_SHEET_NAME]
        header_row = find_pricing_header_row(ws_price)
        fieldnames, rows_iter = iter_table_rows(ws_price, header_row=header_row)
        validate_import_headers(fieldnames)

        def _row_fail(excel_row: int, message: str) -> None:
            stats.failed += 1
            stats.errors.append(f"Row {excel_row}: {message}")
            if strict:
                raise LabPricingImportStrictAbort(
                    f"Row {excel_row}: {message}", stats
                ) from None

        def _process() -> None:
            for excel_row, row in rows_iter:
                avail_raw = row.get("is_available")
                try:
                    if avail_raw in (None, ""):
                        stats.skipped += 1
                        continue
                    availability = parse_bool(avail_raw)
                except ValueError as exc:
                    _row_fail(excel_row, str(exc))
                    continue

                if availability is not True:
                    stats.skipped += 1
                    continue

                err = validate_supported_pricing_row(row)
                if err:
                    _row_fail(excel_row, err)
                    continue

                svc_code = str(row.get("service_code", "")).strip()

                try:
                    svc = _resolve_service(svc_code)
                except DiagnosticServiceMaster.DoesNotExist:
                    _row_fail(excel_row, f"Unknown service_code {svc_code!r}")
                    continue

                sp = parse_decimal(row["selling_price"])
                cp = parse_decimal(row["cost_price"])
                assert sp is not None and cp is not None

                tat_raw = row.get("report_delivery_hours")
                if tat_raw not in (None, ""):
                    report_hours = parse_int(tat_raw) or svc.tat_hours_default
                else:
                    report_hours = svc.tat_hours_default

                hc_raw = row.get("home_collection_supported")
                home_collection = (
                    parse_bool(hc_raw) if hc_raw not in (None, "") else svc.home_collection_possible
                )
                remarks = str(row.get("remarks") or "").strip()

                try:
                    outcome = _upsert_row(
                        branch,
                        svc,
                        selling_price=sp,
                        cost_price=cp,
                        report_hours=report_hours,
                        home_collection=bool(home_collection),
                        remarks=remarks,
                        dry_run=dry_run,
                    )
                except Exception as exc:
                    _row_fail(excel_row, str(exc))
                    continue

                if outcome == "created":
                    stats.created += 1
                elif outcome == "updated":
                    stats.updated += 1
                else:
                    stats.unchanged += 1

        if dry_run:
            _process()
        else:
            with transaction.atomic():
                _process()

        return stats
    finally:
        wb.close()
