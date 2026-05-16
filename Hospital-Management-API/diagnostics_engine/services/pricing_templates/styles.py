from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from diagnostics_engine.services.pricing_templates import constants as C

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=C.HEADER_FILL)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
INSTRUCTION_FONT = Font(bold=True, size=11, color="7F6000")
INSTRUCTION_FILL = PatternFill("solid", fgColor=C.INSTRUCTION_BANNER_FILL)
INSTRUCTION_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

LOCKED_FILL = PatternFill("solid", fgColor=C.LOCKED_COLUMN_FILL)
EDITABLE_FILL = PatternFill("solid", fgColor=C.EDITABLE_COLUMN_FILL)

THIN_SIDE = Side(style="thin", color=C.THIN_BORDER_COLOR)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BOTTOM = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=Side(style="medium", color=C.HEADER_FILL),
)

MISSING_PRICE_FILL = PatternFill("solid", fgColor=C.CONDITIONAL_MISSING_PRICE_FILL)
INVALID_MARGIN_FILL = PatternFill("solid", fgColor=C.CONDITIONAL_INVALID_MARGIN_FILL)
AVAILABLE_TRUE_FILL = PatternFill("solid", fgColor=C.CONDITIONAL_AVAILABLE_TRUE_FILL)
AVAILABLE_FALSE_FILL = PatternFill("solid", fgColor=C.CONDITIONAL_AVAILABLE_FALSE_FILL)
ENABLED_NO_PRICE_FILL = PatternFill("solid", fgColor=C.CONDITIONAL_ENABLED_NO_PRICE_FILL)


def apply_column_widths(ws: Worksheet) -> None:
    for col_letter, width in C.COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].hidden = False
        ws.column_dimensions[col_letter].width = width


def apply_instruction_banner(ws: Worksheet) -> None:
    last_col = get_column_letter(len(C.PRICING_HEADERS))
    ws.merge_cells(f"A{C.ROW_INSTRUCTION}:{last_col}{C.ROW_INSTRUCTION}")
    cell = ws.cell(row=C.ROW_INSTRUCTION, column=1, value=C.INSTRUCTION_BANNER_TEXT)
    cell.font = INSTRUCTION_FONT
    cell.fill = INSTRUCTION_FILL
    cell.alignment = INSTRUCTION_ALIGNMENT
    cell.border = THIN_BORDER


def style_pricing_header_row(ws: Worksheet) -> None:
    for col_idx in range(1, len(C.PRICING_HEADERS) + 1):
        cell = ws.cell(row=C.ROW_HEADER, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BOTTOM


def apply_data_cell_fills(ws: Worksheet, *, max_row: int) -> None:
    """Locked columns = light blue-gray; editable columns = light yellow."""
    if max_row < C.ROW_DATA_START:
        return
    for row_idx in range(C.ROW_DATA_START, max_row + 1):
        for col_idx in range(1, len(C.PRICING_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx in C.EDITABLE_COLUMNS:
                cell.fill = EDITABLE_FILL
            else:
                cell.fill = LOCKED_FILL


def apply_thin_borders(ws: Worksheet, *, max_row: int, max_col: int | None = None) -> None:
    col_count = max_col or len(C.PRICING_HEADERS)
    for row_idx in range(C.ROW_INSTRUCTION, max_row + 1):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == C.ROW_HEADER:
                cell.border = HEADER_BOTTOM
            else:
                cell.border = THIN_BORDER


def apply_conditional_formatting(ws: Worksheet, *, max_row: int) -> None:
    if max_row < C.ROW_DATA_START:
        return
    sp = get_column_letter(C.COL_SELLING_PRICE)
    cp = get_column_letter(C.COL_COST_PRICE)
    avail = get_column_letter(C.COL_IS_AVAILABLE)
    data_start = C.ROW_DATA_START
    last_col = get_column_letter(len(C.PRICING_HEADERS))
    row_range = f"A{data_start}:{last_col}{max_row}"
    margin_range = f"{sp}{data_start}:{cp}{max_row}"

    # Priority: enabled but missing price (red) > available (green) > not offered (gray)
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'AND(${avail}{data_start}="TRUE",ISBLANK(${sp}{data_start}))'],
            fill=ENABLED_NO_PRICE_FILL,
            stopIfTrue=True,
        ),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${avail}{data_start}="TRUE"'],
            fill=AVAILABLE_TRUE_FILL,
        ),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${avail}{data_start}="FALSE"'],
            fill=AVAILABLE_FALSE_FILL,
        ),
    )
    ws.conditional_formatting.add(
        margin_range,
        FormulaRule(
            formula=[
                f'AND(${sp}{data_start}<>"",${cp}{data_start}<>"",${cp}{data_start}>${sp}{data_start})'
            ],
            fill=INVALID_MARGIN_FILL,
        ),
    )


def add_pricing_table(ws: Worksheet, *, max_row: int) -> None:
    """Excel table for filters/stripes — do not also set ws.auto_filter on the same range."""
    if max_row < C.ROW_HEADER:
        return
    last_col = get_column_letter(len(C.PRICING_HEADERS))
    ref = f"A{C.ROW_HEADER}:{last_col}{max_row}"
    table = Table(displayName="PricingCatalog", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_metadata_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True)
            cell.border = THIN_BORDER


def build_instructions_sheet(wb) -> None:
    ws = wb.create_sheet(title=C.INSTRUCTIONS_SHEET_NAME)
    ws.sheet_view.showGridLines = False
    lines = [
        "Lab Pricing Workbook — Quick Guide",
        "",
        "1. Go to the 'pricing_catalog' sheet.",
        "2. Yellow columns are editable. Blue-gray columns are read-only (catalog).",
        "3. Filter column D (lab_department) to work by PATHOLOGY, RADIOLOGY, etc.",
        "4. selling_price: patient price (required for import).",
        "5. cost_price: lab payout; must be <= selling_price.",
        "6. is_available: default FALSE for new labs; set TRUE + price only for offered tests.",
        "7. Rows with blank selling_price are skipped on import.",
        "8. lab_department is for Excel filtering only — not imported to DoctorPro.",
        "",
        "Import: python manage.py sync_lab_pricing --file=<this workbook>",
    ]
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 90
    title = ws["A1"]
    title.font = Font(bold=True, size=14, color="1F4E79")
