from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from diagnostics_engine.services.pricing_templates import constants as C


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def bool_to_excel(value: bool | None, *, default: bool = False) -> str:
    if value is None:
        return C.BOOLEAN_TRUE if default else C.BOOLEAN_FALSE
    return C.BOOLEAN_TRUE if value else C.BOOLEAN_FALSE


def read_metadata_kv(ws: Worksheet) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        key = row[0]
        if key is None or str(key).strip() == "":
            continue
        val = row[1] if len(row) > 1 else None
        out[str(key).strip()] = "" if val is None else str(val).strip()
    return out


def read_metadata_sheet(ws: Worksheet) -> Dict[str, str]:
    """Alias for workbook metadata Key/Value sheet (same as read_metadata_kv)."""
    return read_metadata_kv(ws)


def validate_import_metadata(meta: Dict[str, str]) -> None:
    if not (meta.get("branch_code") or "").strip():
        raise ValueError("metadata sheet: branch_code is required")
    if not (meta.get("template_version") or "").strip():
        raise ValueError("metadata sheet: template_version is required")


def find_pricing_header_row(ws: Worksheet) -> int:
    """Locate header row (supports instruction banner above headers)."""
    max_scan = min(ws.max_row or 1, 200)
    for row_idx in range(1, max_scan + 1):
        val = ws.cell(row=row_idx, column=1).value
        if normalize_header(val) == "service_code":
            return row_idx
    return C.ROW_HEADER


def iter_table_rows(
    ws: Worksheet,
    header_row: int | None = None,
) -> Tuple[List[str], Iterable[Tuple[int, Dict[str, Any]]]]:
    header_row = header_row or find_pricing_header_row(ws)
    header_cells_list = list(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    header = header_cells_list[0] if header_cells_list else None
    if not header:
        return [], iter(())

    fieldnames = [normalize_header(h) for h in header]
    ncol = len(header)

    def _gen():
        for row in ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ncol,
        ):
            excel_row = row[0].row
            raw = tuple(c.value for c in row)
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            row_dict: Dict[str, Any] = {}
            for key, val in zip(fieldnames, raw):
                if not key:
                    continue
                row_dict[key] = val
            yield excel_row, row_dict

    return fieldnames, _gen()


def parse_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().upper()
    if text in {C.BOOLEAN_TRUE, "T", "YES", "Y", "1"}:
        return True
    if text in {C.BOOLEAN_FALSE, "F", "NO", "N", "0"}:
        return False
    raise ValueError(f"Invalid boolean: {value!r}")


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal: {value!r}") from exc


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"Invalid integer: {value!r}") from exc
