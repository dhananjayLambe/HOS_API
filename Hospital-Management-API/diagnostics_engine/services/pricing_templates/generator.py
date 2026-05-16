from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection

from diagnostics_engine.models.catalog import DiagnosticServiceMaster
from diagnostics_engine.services.pricing_templates import constants as C
from diagnostics_engine.services.pricing_templates.constants import resolve_lab_department
from diagnostics_engine.services.pricing_templates.excel_utils import bool_to_excel
from diagnostics_engine.services.pricing_templates.styles import (
    add_pricing_table,
    apply_column_widths,
    apply_conditional_formatting,
    apply_data_cell_fills,
    apply_instruction_banner,
    apply_thin_borders,
    build_instructions_sheet,
    style_metadata_sheet,
    style_pricing_header_row,
)
from diagnostics_engine.services.pricing_templates.validators import (
    add_boolean_dropdown,
    add_positive_decimal_validation,
    add_tat_integer_validation,
)
from labs.models import BranchServicePricing, LabBranch


def _branch_metadata(branch: LabBranch, generated_by: str) -> Dict[str, str]:
    try:
        address = branch.address
    except ObjectDoesNotExist:
        address = None
    city = address.city if address else ""
    pincode = address.pincode if address else ""
    return {
        "branch_code": branch.branch_code,
        "lab_name": branch.organization.organization_name,
        "branch_name": branch.branch_name,
        "city": city,
        "pincode": pincode,
        "template_version": C.TEMPLATE_VERSION,
        "generated_at": timezone.now().isoformat(timespec="seconds"),
        "generated_by": generated_by,
    }


def _branch_pricing_map(branch: LabBranch) -> Dict[str, BranchServicePricing]:
    out: Dict[str, BranchServicePricing] = {}
    for pricing in BranchServicePricing.objects.filter(
        branch=branch,
        is_deleted=False,
        is_active=True,
    ).select_related("service"):
        out[pricing.service.code] = pricing
    return out


def _sorted_services() -> List[DiagnosticServiceMaster]:
    services = list(
        DiagnosticServiceMaster.objects.filter(is_active=True, deleted_at__isnull=True).select_related(
            "category", "category__parent"
        )
    )

    def sort_key(svc: DiagnosticServiceMaster) -> tuple:
        parent_name = svc.category.parent.name if svc.category.parent_id else None
        dept = resolve_lab_department(svc.category.name, parent_name=parent_name)
        return (dept, svc.category.name.lower(), svc.name.lower())

    services.sort(key=sort_key)
    return services


def _apply_sheet_protection(ws, *, max_row: int) -> None:
    for row in ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=len(C.PRICING_HEADERS),
    ):
        for cell in row:
            if cell.row < C.ROW_DATA_START:
                locked = True
            else:
                locked = cell.column not in C.EDITABLE_COLUMNS
            cell.protection = Protection(locked=locked)

    prot = ws.protection
    prot.sheet = True
    prot.autoFilter = False
    prot.sort = False
    prot.formatCells = False
    prot.formatColumns = False
    prot.formatRows = False
    prot.insertRows = False
    prot.deleteRows = False
    prot.insertColumns = True
    prot.deleteColumns = True
    prot.selectLockedCells = False
    prot.selectUnlockedCells = False
    prot.enable()


def _apply_metadata_protection(ws) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.enable()


def build_lab_pricing_workbook(branch: LabBranch, *, generated_by: str | None = None) -> Workbook:
    wb = Workbook()
    wb.security = WorkbookProtection(lockStructure=True, lockWindows=False)

    ws_meta = wb.active
    ws_meta.title = C.METADATA_SHEET_NAME
    ws_meta.append(["Key", "Value"])
    meta = _branch_metadata(branch, generated_by or C.GENERATED_BY_DEFAULT)
    for key in C.METADATA_KEYS:
        ws_meta.append([key, meta.get(key, "")])
    ws_meta.append(
        [
            "pricing_catalog_hint",
            "See 'pricing_catalog' (column D = lab_department filter) and 'instructions' sheet.",
        ]
    )
    style_metadata_sheet(ws_meta)
    _apply_metadata_protection(ws_meta)

    build_instructions_sheet(wb)

    ws = wb.create_sheet(title=C.PRICING_SHEET_NAME)
    apply_instruction_banner(ws)
    for col_idx, header in enumerate(C.PRICING_HEADERS, start=1):
        ws.cell(row=C.ROW_HEADER, column=col_idx, value=header)
    style_pricing_header_row(ws)

    pricing_by_code = _branch_pricing_map(branch)

    for svc in _sorted_services():
        pricing = pricing_by_code.get(svc.code)
        remarks = ""
        if pricing and isinstance(pricing.metadata, dict):
            remarks = str(pricing.metadata.get("remarks") or "")

        parent_name = svc.category.parent.name if svc.category.parent_id else None
        lab_department = resolve_lab_department(svc.category.name, parent_name=parent_name)

        home_default = svc.home_collection_possible
        report_hours = svc.tat_hours_default
        selling_price = None
        cost_price = None
        is_available_excel = C.BOOLEAN_FALSE

        if pricing:
            selling_price = pricing.selling_price
            cost_price = pricing.cost_price
            report_hours = pricing.report_delivery_hours
            home_default = pricing.home_collection_supported
            is_available_excel = bool_to_excel(pricing.is_available)

        ws.append(
            [
                svc.code,
                svc.name,
                svc.category.name,
                lab_department,
                svc.sample_type or "",
                svc.tat_hours_default,
                selling_price,
                cost_price,
                report_hours,
                bool_to_excel(home_default, default=svc.home_collection_possible),
                is_available_excel,
                remarks,
            ]
        )

    max_row = ws.max_row
    ws.freeze_panes = C.FREEZE_PANES

    if max_row >= C.ROW_DATA_START:
        start_row, end_row = C.ROW_DATA_START, max_row
        add_boolean_dropdown(
            ws, get_column_letter(C.COL_HOME_COLLECTION), start_row, end_row
        )
        add_boolean_dropdown(
            ws,
            get_column_letter(C.COL_IS_AVAILABLE),
            start_row,
            end_row,
            false_first=True,
            allow_blank=True,
        )
        add_positive_decimal_validation(
            ws, get_column_letter(C.COL_SELLING_PRICE), start_row, end_row
        )
        add_positive_decimal_validation(
            ws, get_column_letter(C.COL_COST_PRICE), start_row, end_row
        )
        add_tat_integer_validation(
            ws, get_column_letter(C.COL_REPORT_TAT), start_row, end_row
        )

    apply_column_widths(ws)
    apply_data_cell_fills(ws, max_row=max_row)
    apply_thin_borders(ws, max_row=max_row)
    apply_conditional_formatting(ws, max_row=max_row)
    add_pricing_table(ws, max_row=max_row)
    _apply_sheet_protection(ws, max_row=max_row)

    wb.active = wb[C.PRICING_SHEET_NAME]
    return wb


def save_lab_pricing_workbook(
    branch: LabBranch,
    wb: Workbook,
    *,
    output_dir: Path | None = None,
    force: bool = False,
) -> Path:
    base_dir = Path(output_dir) if output_dir else Path(settings.MEDIA_ROOT) / "lab_pricing_templates"
    base_dir.mkdir(parents=True, exist_ok=True)
    filename = f"LabPricing_{branch.branch_code}_{C.TEMPLATE_VERSION}.xlsx"
    path = base_dir / filename
    if path.exists() and not force:
        raise FileExistsError(f"Template already exists: {path} (use force=True to overwrite)")
    wb.save(str(path))
    return path
