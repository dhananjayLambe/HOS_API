import csv
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from medicines.models import DrugMaster, FormulationMaster


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


_STRENGTH_RE = re.compile(
    r"\d+(?:\.\d+)?\s*(?:mg|mcg|g|ml|%|w/v)"
    r"(?:\s*/\s*\d+(?:\.\d+)?\s*(?:mg|mcg|g|ml|%|w/v))*",
    re.IGNORECASE,
)


def _first_value(row, *keys):
    for key in keys:
        raw = row.get(key)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            return text
    return ""


def _infer_strength_from_label(text):
    if not text:
        return ""
    m = _STRENGTH_RE.search(text)
    return m.group(0).replace(" ", "") if m else ""


def _iter_rows_csv(path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    yield {_normalize_header(k): (v or "").strip() for k, v in row.items()}
            return
        except UnicodeDecodeError as e:
            last_error = e
            continue
    raise CommandError(f"Could not decode CSV (tried utf-8, cp1252, latin-1): {last_error}")


def _iter_rows_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return
        fieldnames = [_normalize_header(h) for h in header]
        for row in rows:
            if row is None:
                continue
            cells = list(row)
            if all(c is None or str(c).strip() == "" for c in cells):
                continue
            d = {}
            for key, val in zip(fieldnames, cells):
                if not key:
                    continue
                d[key] = "" if val is None else str(val).strip()
            yield d
    finally:
        wb.close()


def iter_medicine_rows(file_path):
    path = Path(file_path)
    if not path.is_file():
        raise CommandError(f"File not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        yield from _iter_rows_xlsx(path)
    elif suffix == ".xls":
        raise CommandError("Legacy .xls is not supported; save the sheet as .xlsx and retry.")
    else:
        yield from _iter_rows_csv(path)


class Command(BaseCommand):
    help = "Import medicines from a CSV or Excel (.xlsx) file."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"]
        created = 0
        updated = 0

        with transaction.atomic():
            for row in iter_medicine_rows(file_path):
                brand_name = _first_value(row, "brand_name", "product_name")
                generic_name = _first_value(
                    row, "generic_name", "salt_composition", "composition"
                )
                formulation_name = _first_value(
                    row, "formulation", "product_form", "medicine_type"
                )
                manufacturer = _first_value(
                    row, "manufacturer", "marketer/_manufacturer"
                )
                strength = _first_value(row, "strength")
                if not strength:
                    strength = _infer_strength_from_label(brand_name)

                if not brand_name or not formulation_name:
                    continue

                formulation, _ = FormulationMaster.objects.get_or_create(
                    name=formulation_name.lower()
                )

                explicit_code = _first_value(row, "code", "product_id")
                if explicit_code:
                    code = explicit_code[:50]
                else:
                    code = f"{brand_name[:10].upper()}-{strength}".replace(" ", "")
                    code = code[:50]

                drug, is_created = DrugMaster.objects.update_or_create(
                    code=code,
                    defaults={
                        "brand_name": brand_name,
                        "generic_name": generic_name,
                        "strength": strength,
                        "formulation": formulation,
                        "manufacturer": manufacturer,
                        "is_active": True,
                    },
                )

                if is_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import completed: {created} created, {updated} updated"
            )
        )
