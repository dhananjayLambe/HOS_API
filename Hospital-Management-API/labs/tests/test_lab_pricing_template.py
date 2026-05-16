"""Tests for lab pricing template generation and sync."""

from __future__ import annotations

import tempfile
import uuid
from decimal import Decimal
from pathlib import Path

from django.test import TestCase, override_settings
from openpyxl import load_workbook

from diagnostics_engine.models.catalog import DiagnosticCategory, DiagnosticServiceMaster
from diagnostics_engine.models.choices import CommissionType
from diagnostics_engine.services.pricing_templates import constants as C
from diagnostics_engine.services.pricing_templates.constants import resolve_lab_department
from diagnostics_engine.services.pricing_templates.excel_utils import find_pricing_header_row
from diagnostics_engine.services.pricing_templates.generator import (
    build_lab_pricing_workbook,
    save_lab_pricing_workbook,
)
from diagnostics_engine.services.pricing_templates.importer import import_lab_pricing
from labs.choices.auth import LabType, RegistrationStatus
from labs.models import BranchServicePricing, LabAddress, LabBranch, LabOrganization


def _make_branch_with_services(
    *,
    service_count: int = 2,
    category_name: str | None = None,
) -> tuple[LabBranch, list[DiagnosticServiceMaster]]:
    suffix = uuid.uuid4().hex[:8]
    org = LabOrganization.objects.create(
        organization_name=f"Test Lab {suffix}",
        display_name=f"Test Lab {suffix}",
        organization_code=f"ORG-{suffix}",
        slug=f"test-lab-{suffix}",
        lab_type=LabType.PATHOLOGY_LAB,
        owner_name="Owner",
        primary_contact_number="9999999999",
        registration_status=RegistrationStatus.APPROVED,
        is_verified=True,
        onboarding_completed=True,
        is_active_for_orders=True,
    )
    branch = LabBranch.objects.create(
        organization=org,
        branch_name="Main Branch",
        branch_code=f"BR-{suffix}",
        is_active=True,
        is_active_for_orders=True,
    )
    LabAddress.objects.create(
        branch=branch,
        address_line_1="1 Test St",
        city="Pune",
        state="Maharashtra",
        pincode="412207",
    )
    cat = DiagnosticCategory.objects.create(
        name=category_name or f"Category {suffix}",
        code=f"CAT-{suffix}",
    )
    services = []
    for i in range(service_count):
        svc = DiagnosticServiceMaster.objects.create(
            code=f"SVC-{suffix}-{i}",
            name=f"Service {i}",
            category=cat,
            sample_type="Blood",
            tat_hours_default=24,
            home_collection_possible=True,
        )
        services.append(svc)
    return branch, services


class LabPricingTemplateTests(TestCase):
    def test_resolve_lab_department_mapping(self):
        self.assertEqual(resolve_lab_department("Hematology"), "PATHOLOGY")
        self.assertEqual(resolve_lab_department("X-Ray"), "RADIOLOGY")
        self.assertEqual(resolve_lab_department("ECG", parent_name="Cardiology"), "CARDIOLOGY")
        self.assertEqual(resolve_lab_department("Unknown Category XYZ"), "OTHER")

    def test_workbook_generation(self):
        branch, services = _make_branch_with_services(service_count=2)
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                wb = build_lab_pricing_workbook(branch)
                path = save_lab_pricing_workbook(branch, wb, force=True)

            self.assertTrue(path.exists())
            loaded = load_workbook(path, data_only=True)
            self.assertIn(C.METADATA_SHEET_NAME, loaded.sheetnames)
            self.assertIn(C.PRICING_SHEET_NAME, loaded.sheetnames)
            self.assertIn(C.INSTRUCTIONS_SHEET_NAME, loaded.sheetnames)

            ws = loaded[C.PRICING_SHEET_NAME]
            header_row = find_pricing_header_row(ws)
            self.assertEqual(header_row, C.ROW_HEADER)
            headers = [ws.cell(header_row, c).value for c in range(1, len(C.PRICING_HEADERS) + 1)]
            self.assertEqual(headers, list(C.PRICING_HEADERS))
            self.assertEqual(headers[3], "lab_department")
            first_data_row = C.ROW_DATA_START
            self.assertEqual(ws.max_row - first_data_row + 1, len(services))
            self.assertEqual(ws.cell(first_data_row, 1).value, services[0].code)
            self.assertEqual(ws.freeze_panes, C.FREEZE_PANES)
            loaded.close()

    def test_lab_department_column_for_hematology(self):
        branch, _services = _make_branch_with_services(service_count=1, category_name="Hematology")
        wb = build_lab_pricing_workbook(branch)
        ws = wb[C.PRICING_SHEET_NAME]
        self.assertEqual(ws.cell(C.ROW_DATA_START, C.COL_LAB_DEPARTMENT).value, "PATHOLOGY")

    def test_default_is_available_false_without_pricing(self):
        branch, _services = _make_branch_with_services(service_count=1)
        wb = build_lab_pricing_workbook(branch)
        ws = wb[C.PRICING_SHEET_NAME]
        r = C.ROW_DATA_START
        self.assertEqual(ws.cell(r, C.COL_IS_AVAILABLE).value, C.BOOLEAN_FALSE)
        self.assertIsNone(ws.cell(r, C.COL_SELLING_PRICE).value)
        self.assertIsNone(ws.cell(r, C.COL_COST_PRICE).value)

    def test_prefill_existing_pricing(self):
        branch, services = _make_branch_with_services(service_count=1)
        svc = services[0]
        BranchServicePricing.objects.create(
            branch=branch,
            service=svc,
            selling_price=Decimal("499.00"),
            cost_price=Decimal("350.00"),
            platform_margin_type=CommissionType.FLAT,
            platform_margin_value=Decimal("0"),
            doctor_commission_type=CommissionType.FLAT,
            doctor_commission_value=Decimal("0"),
            report_delivery_hours=48,
            home_collection_supported=True,
            is_available=True,
            metadata={"remarks": "VIP rate"},
        )

        wb = build_lab_pricing_workbook(branch)
        ws = wb[C.PRICING_SHEET_NAME]
        r = C.ROW_DATA_START
        self.assertEqual(ws.cell(r, C.COL_SELLING_PRICE).value, Decimal("499.00"))
        self.assertEqual(ws.cell(r, C.COL_COST_PRICE).value, Decimal("350.00"))
        self.assertEqual(ws.cell(r, C.COL_REPORT_TAT).value, 48)
        self.assertEqual(ws.cell(r, C.COL_REMARKS).value, "VIP rate")
        self.assertEqual(ws.cell(r, C.COL_IS_AVAILABLE).value, C.BOOLEAN_TRUE)

    def test_import_updates_db(self):
        branch, services = _make_branch_with_services(service_count=1)
        svc = services[0]

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                wb = build_lab_pricing_workbook(branch)
                ws = wb[C.PRICING_SHEET_NAME]
                r = C.ROW_DATA_START
                ws.cell(r, C.COL_SELLING_PRICE).value = 599
                ws.cell(r, C.COL_COST_PRICE).value = 400
                ws.cell(r, C.COL_REPORT_TAT).value = 36
                ws.cell(r, C.COL_IS_AVAILABLE).value = C.BOOLEAN_TRUE
                path = save_lab_pricing_workbook(branch, wb, force=True)

            stats = import_lab_pricing(path)
            self.assertEqual(stats.created, 1)
            self.assertEqual(stats.failed, 0)

            pricing = BranchServicePricing.objects.get(branch=branch, service=svc, is_active=True)
            self.assertEqual(pricing.selling_price, Decimal("599"))
            self.assertEqual(pricing.cost_price, Decimal("400"))
            self.assertEqual(pricing.report_delivery_hours, 36)

    def test_lab_department_ignored_on_import(self):
        branch, services = _make_branch_with_services(service_count=1, category_name="Hematology")
        svc = services[0]

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                wb = build_lab_pricing_workbook(branch)
                ws = wb[C.PRICING_SHEET_NAME]
                r = C.ROW_DATA_START
                ws.cell(r, C.COL_LAB_DEPARTMENT).value = "RADIOLOGY"
                ws.cell(r, C.COL_SELLING_PRICE).value = 150
                ws.cell(r, C.COL_COST_PRICE).value = 100
                ws.cell(r, C.COL_IS_AVAILABLE).value = C.BOOLEAN_TRUE
                path = save_lab_pricing_workbook(branch, wb, force=True)

            stats = import_lab_pricing(path)
            self.assertEqual(stats.created, 1)
            pricing = BranchServicePricing.objects.get(branch=branch, service=svc, is_active=True)
            self.assertEqual(pricing.selling_price, Decimal("150"))

    def test_blank_selling_price_skipped(self):
        branch, services = _make_branch_with_services(service_count=1)

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                wb = build_lab_pricing_workbook(branch)
                ws = wb[C.PRICING_SHEET_NAME]
                r = C.ROW_DATA_START
                ws.cell(r, C.COL_SELLING_PRICE).value = None
                ws.cell(r, C.COL_COST_PRICE).value = None
                path = save_lab_pricing_workbook(branch, wb, force=True)

            stats = import_lab_pricing(path)
            self.assertEqual(stats.skipped, 1)
            self.assertEqual(stats.created, 0)
            self.assertFalse(
                BranchServicePricing.objects.filter(branch=branch, service=services[0]).exists()
            )

    def test_invalid_service_code_fails(self):
        branch, _services = _make_branch_with_services(service_count=1)

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                wb = build_lab_pricing_workbook(branch)
                ws = wb[C.PRICING_SHEET_NAME]
                r = C.ROW_DATA_START
                ws.cell(r, C.COL_SERVICE_CODE).value = "UNKNOWN-SVC"
                ws.cell(r, C.COL_SELLING_PRICE).value = 100
                ws.cell(r, C.COL_COST_PRICE).value = 80
                ws.cell(r, C.COL_IS_AVAILABLE).value = C.BOOLEAN_TRUE
                path = save_lab_pricing_workbook(branch, wb, force=True)

            stats = import_lab_pricing(path)
            self.assertEqual(stats.failed, 1)
            self.assertTrue(any("Unknown service_code" in e or "unknown" in e.lower() for e in stats.errors))
